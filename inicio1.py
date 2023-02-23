import openpyxl
from getpass import getuser
from datetime import datetime
from tkinter import *
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
import pythoncom
import os
import win32com.client
from reporte import ingresarsap,meteteensap
import shutil
import threading
import queue



def copiar_rechazar():
    ######INTERFAZ GRAFICA#######
    user=getuser()
    root=Tk()
    dt = datetime.now()
    day = dt.day
    month = dt.month
    year = dt.year
    hoy = str(day) + "-" + str(month) + "-" + str(year)
    root.title("Copiar y Rechazar OyP-Sistemas")
    root.resizable(0,0)
    root.geometry('450x250+450+200'.format(400, 400))
    miFrame = Frame(root, width=1000)
    miFrame.pack()
    miFrame2 = Frame(root, width=1000)
    miFrame2.pack()
    miFrame3 = Frame(root, width=1000)
    miFrame3.pack()
    miFrame4 = Frame(root, width=1000)
    miFrame4.pack()

    ######CAMPOS DE USUARIO#######
    usuariosap = StringVar()
    cliente1 = Entry(miFrame,textvariable=usuariosap,width=20)
    cliente1.grid(row=3,column=1,padx=30,pady=5)
    rutalabel2 = Label(miFrame,text="Usuario",fg="Black",font=('Bold 1',10))
    rutalabel2.grid(row=2,column=1,sticky="n",padx=0,pady=0)
    
    ######CAMPOS DE CONTRASEÑA#######
    contrasegna = StringVar()
    cliente1 = Entry(miFrame,textvariable=contrasegna,width=20,show="*")
    cliente1.grid(row=3,column=3,padx=30,pady=5)
    rutalabel2 = Label(miFrame,text="Contraseña",fg="Black",font=('Bold 1',10))
    rutalabel2.grid(row=2,column=3,sticky="n",padx=0,pady=0)

    
    def leerexcel():
        ######LEER EL EXCEL#######
        pathagenda = "C:/Users/"+ user + "/Desktop/Copiar_Rechazar/pedidos.xlsx"
        wb = openpyxl.load_workbook(pathagenda,data_only=True)
        ws = wb["DETALLE"]
        ultimafiladelws = len(ws['A'])
        pedidos = []
        fila = []
        
        
        ######TOMAR LOS DATOS DEL EXCEL#######
        for dato in range(2, ultimafiladelws + 1):
            validacion = ws.cell(row = dato,column = 1).value
            if validacion == None or validacion == "":
                continue
            else:
                pedidos.append(ws.cell(row = dato, column = 1).value)
                fila.append(ws.cell(row = dato, column = 3).internal_value)
                
        wb.close()
        return pedidos,fila
    
    def procesarensap(pedidos, fila, sesion, grupo):    
        pythoncom.CoInitialize()
        print(grupo)
        generarexcel(grupo)
        pythoncom.CoInitialize()
        path = "C:/Users/"+ user + "/Desktop/Copiar_Rechazar/pedidos_grupo" + str(grupo) +".xlsx"
        wb = openpyxl.load_workbook(path)
        ws = wb["DETALLE"]

        cantidaddeped = len(pedidos)
        for i in range(2,5000):
            ws.cell(row=i,column=1).value = ""
            ws.cell(row=i,column=2).value = ""
            ws.cell(row=i,column=3).value = ""

        resultado = []
        for ped in pedidos:
            res = meteteensap(ped,sesion)
            resultado.append(res)

        cantidadderesultado = len(resultado)
        for i in range(0, cantidadderesultado):
            filanueva = i + 2
            if pedidos[i] == "" or pedidos[i] == None:
                continue
            else:
                ws.cell(row=filanueva,column=1).value=pedidos[i]
                ws.cell(row=filanueva,column=2).value=resultado[i]
                ws.cell(row=filanueva,column=3).value=fila[i]
        wb.save(path)
        wb.close()
        

        
    def generarexcel(grupo):
        base=("C:/Users/"+ user + "/Desktop/Copiar_Rechazar/pedidos.xlsx")
        shutil.copy(base,"C:/Users/"+ user + "/Desktop/Copiar_Rechazar/pedidos_grupo" + str(grupo) +".xlsx")
    
    def control(grupo):
        ######CARGAR EXCEL#######
        res_pedido=[]
        res_fila=[]
        res_resultado=[]
        for j in range(0,grupo):
            pathagenda="C:/Users/"+ user + "/Desktop/Copiar_Rechazar/pedidos_grupo" + str(j) +".xlsx"
            wb = openpyxl.load_workbook(pathagenda)
            ws = wb["DETALLE"]
            cantidadenetregas=len(ws['A'])
            for i in range(2,cantidadenetregas):
                if ws.cell(row=i,column=1).value=="" or ws.cell(row=i,column=1).value==None:
                    continue
                else:
                    res_pedido.append(ws.cell(row=i,column=1).value)
                    res_fila.append(ws.cell(row=i,column=3).value)
                    res_resultado.append(ws.cell(row=i,column=2).value)
        for j in range(0,grupo):    
            os.remove("C:/Users/"+ user + "/Desktop/Copiar_Rechazar/pedidos_grupo" + str(j) +".xlsx")
        pathagenda="C:/Users/"+ user + "/Desktop/Copiar_Rechazar/pedidos.xlsx"
        wb=openpyxl.load_workbook(pathagenda)
        ws = wb["DETALLE"]
        cantidadenetregas=len(ws['A'])
        print(res_pedido)
        print(res_resultado)
        print(len(res_pedido))
        for i in range(0,len(res_pedido)):
            try:
                if (ws.cell(row=res_fila[i],column=1).value)==res_pedido[i]:
                    ws.cell(row=res_fila[i],column=2).value=res_resultado[i]
            except:
                break
        wb.save("C:/Users/"+ user + "/Desktop/Copiar_Rechazar/pedidos.xlsx")
        wb.close()



    def proceso():
        ingresarsap(usuariosap.get(), contrasegna.get())
        pythoncom.CoInitialize()
        #######COPIAR Y RECHAZAR POR PEDIDO EN SAP
        pathagenda="C:/Users/"+ user + "/Desktop/Copiar_Rechazar/pedidos.xlsx"
        wb = openpyxl.load_workbook(pathagenda,data_only=True)
        ws = wb["DETALLE"]
        pedidos=[]
        fila=[]
        
        intervalo1=[]
        intervalo2=[]
        intervalo3=[]
        intervalo4=[]
        intervalo5=[]
        intervalo6=[]
        intervalo7=[]
        intervalo8=[]
        intervalo9=[]
        intervalo10=[]
        intervalo11=[]
        intervalo12=[]
        intervalo13=[]
        intervalo14=[]
        intervalo15=[]
        intervalo16=[]
        intervalo17=[]
        intervalo18=[]
        grupo=[]
        
        ultimafiladelws = len(ws['A'])
        
        for dato in range(2, ultimafiladelws + 1):
            if dato == None:
                continue
            else:
                pedidos.append(ws.cell(row=dato,column=1).value)
                fila.append(ws.cell(row=dato,column=3).internal_value)
        alcance = len(pedidos)
        
        for i in range(0, alcance + 1):
            if i<40:
                intervalo1.append(i)
            elif i<80:
                intervalo2.append(i)
            elif i<120:
                intervalo3.append(i)
            elif i<160:
                intervalo4.append(i)
            elif i<500:
                intervalo5.append(i)

            
        listadegrupos=[intervalo1,intervalo2,intervalo3,intervalo4,intervalo5,intervalo6]
        
        for i in listadegrupos:
            if i!=[]:
                grupo.append(i)
        
        cantidaddegrupos=len(grupo)
        print(cantidaddegrupos)
        
        wb.close()
        hilos=[]
        NUM_HILOS2=cantidaddegrupos
        for num_hilo in range(NUM_HILOS2):
            inter=grupo[num_hilo]
            posicion0=inter[0]
            sesion=num_hilo
            posicionn=inter[-1]
            corte_pedidos=pedidos[posicion0:posicionn+1]
            fila_corte=fila[posicion0:posicionn+1]
            hilo = threading.Thread(name='hilo_remito%s' %num_hilo,target=procesarensap,
                                    args=(corte_pedidos,fila_corte,sesion,num_hilo,),daemon=False)
            hilos.append(hilo)
            hilo.start()
        hiloprincipal=threading.main_thread()
        for i in hilos:
            while i.is_alive()==True:
                continue
        control(cantidaddegrupos)
        tk.messagebox.showinfo(title='Proceso terminado', message=usuariosap.get() + ' el proceso finalizó con exito.')
    botonarmado=Button(miFrame3, text="Inicio",command=proceso)
    botonarmado.grid(row=5,column=2,sticky="e",padx=10,pady=10)


    root.mainloop()
    
#---------------------------------------------------------fin-----------------------------------------------------------#
if __name__=="__main__":
    copiar_rechazar()   