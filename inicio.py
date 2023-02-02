from lectura_xml import lector
from enviar_mail import enviarmail_sin_error
from enviar_mail import enviarmail_con_error
import openpyxl
from getpass import getuser
from datetime import datetime
from tkinter import *
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
import pythoncom
import os
import win32com.client
from reporte import ingresarsap,meteteensap
import shutil



def doka_oyp():
    ######INTERFAZ GRAFICA#######
    user=getuser()
    root=Tk()
    dt = datetime.now()
    day = dt.day
    month = dt.month
    year = dt.year
    hoy=str(day) + "-" + str(month) + "-" + str(year)
    root.title("Doka_Mail")
    root.resizable(0,0)
    root.geometry('450x250+450+200'.format(400, 400))
    miFrame=Frame(root,width=1000)
    miFrame.pack()
    miFrame2=Frame(root,width=1000)
    miFrame2.pack()
    miFrame3=Frame(root,width=1000)
    miFrame3.pack()
    miFrame4=Frame(root,width=1000)
    miFrame4.pack()

    
    ######CAMPOS DE INTERFAZ#######
    id0=IntVar()
    cliente0=Entry(miFrame4,textvariable=id0,width=20)
    cliente0.grid(row=2,column=2,padx=10,pady=5)
    rutalabel1=Label(miFrame4,text="Fecha DOKA (AAMMDD)",fg="Black",font=('Bold 1',10))
    rutalabel1.grid(row=1,column=2,sticky="n",padx=0,pady=0)
    
    
        ######CAMPOS DE INTERFAZ#######
    usuariosap=StringVar()
    cliente1=Entry(miFrame,textvariable=usuariosap,width=20)
    cliente1.grid(row=3,column=1,padx=30,pady=5)
    rutalabel2=Label(miFrame,text="Usuario",fg="Black",font=('Bold 1',10))
    rutalabel2.grid(row=2,column=1,sticky="n",padx=0,pady=0)
    
    
    contrasegna=StringVar()
    cliente1=Entry(miFrame,textvariable=contrasegna,width=20,show="*")
    cliente1.grid(row=3,column=3,padx=30,pady=5)
    rutalabel2=Label(miFrame,text="Contraseña",fg="Black",font=('Bold 1',10))
    rutalabel2.grid(row=2,column=3,sticky="n",padx=0,pady=0)

    hora=StringVar()
    cliente1=Entry(miFrame2,textvariable=hora,width=20)
    cliente1.grid(row=4,column=5,padx=30,pady=5)
    rutalabel2=Label(miFrame2,text="Hora_Corte_SAP(23,09,13)",fg="Black",font=('Bold 1',10))
    rutalabel2.grid(row=3,column=5,sticky="n",padx=0,pady=0)
    
    fecha_corte=StringVar()
    cliente1=Entry(miFrame2,textvariable=fecha_corte,width=20)
    cliente1.grid(row=4,column=1,padx=30,pady=5)
    rutalabel2=Label(miFrame2,text="Fecha_Corte_SAP(23,09,13)",fg="Black",font=('Bold 1',10))
    rutalabel2.grid(row=3,column=1,sticky="n",padx=0,pady=0)
    
    
    def leerexcel():
        ######LEER EL EXCEL#######
        pathagenda="C:/Users/"+ user + "/Desktop/doka/doka_oyp.xlsm"
        wb = openpyxl.load_workbook(pathagenda,data_only=True,keep_vba=True)
        ws = wb["DETALLE ENVÍO DE MAIL"]
        ultimafiladelws=len(ws['A'])
        entregas=[]
        mail=[]
        nombre=[]
        asunto=[]
        texto=[]
        farmacia=[]
        resultado=[]
        fila=[]
        
        
        ######TOMAR LOS DATOS DEL EXCEL#######
        for dato in range(2,ultimafiladelws+1):
            if dato==None:
                continue
            else:
                entregas.append(ws.cell(row=dato,column=1).value)
                mail.append(ws.cell(row=dato,column=2).value)
                nombre.append(ws.cell(row=dato,column=3).value)
                farmacia.append(ws.cell(row=dato,column=4).value)
                asunto.append(ws.cell(row=dato,column=5).internal_value)
                texto.append(ws.cell(row=dato,column=6).value)
                fila.append(ws.cell(row=dato,column=7).value)
                resultado.append(ws.cell(row=dato,column=8).value)
                
        wb.close()
        return entregas,mail,nombre,farmacia,asunto,texto,fila,resultado
    
    def generarexcel():
        base=("C:/Users/"+ user + "/Desktop/doka/doka_oyp_base.xlsm")
        shutil.copy(base,"C:/Users/"+ user + "/Desktop/doka/doka_oyp.xlsm")

    def cargardatos():
        
        generarexcel()
        
        ingresarsap(usuariosap.get(),contrasegna.get())
        
        meteteensap(str(fecha_corte.get()),str(hora.get()),0)
            
        pythoncom.CoInitialize()
        path="C:/Users/"+ user + "/Desktop/doka/doka_oyp.xlsm"
        if os.path.exists(path):
            pythoncom.CoInitialize()
            Excel_macro = win32com.client.DispatchEx("Excel.Application") # DispatchEx is required in the newest versions of Python.
            Excel_path = os.path.expanduser(path)
            workbook = Excel_macro.Workbooks.Open(Filename = Excel_path, ReadOnly =1)
            Excel_macro.Application.Run("doka_oyp.xlsm" + "!" + "z_integrador_doka.facturacion") # update Module1 with your module, Macro1 with your macro
            workbook.Save()
            workbook.Close()
            Excel_macro.Application.Quit()  
            del Excel_macro
            pass
        
    
    
    def mail():
        
        ######CAMPOS DE INTERFAZ#######
        rutahtml="C:/Users/" + user + "/Desktop/doka/ticket.html"
        dia=id0.get()
        
        
        ######UNPACKING DE COLUMNAS#######
        entregas,mail,nombre,farmacia,asunto,texto,fila,resultado=leerexcel()  
        cantidadenetregas=len(entregas)
        
        ######CARGAR EXCEL#######
        pathagenda="C:/Users/"+ user + "/Desktop/doka/doka_oyp.xlsm"
        wb = openpyxl.load_workbook(pathagenda,keep_vba=True)
        ws = wb["DETALLE"]
        
        
        ######PROCESO POR LINEA#######
        for i in range(0,cantidadenetregas):
            if mail[i]=="NOSUSAR@SCIENZA.COM.AR" or mail[i]=="" or mail[i]==None or mail[i]=="NOSUSAR@SCIENZA.COM" or mail[i]==" ": ######CONDICIONES DE ERROR DEL MAIL#######
                
                ######REEMPLAZO DE VALORES A EXCEPCIPON DE ERROR#######
                mail[i]="serviciopostentrega@scienza.com.ar"
                asunto[i]="TICKET NO ENVIADO POR ERROR EN LOS DATOS DEL AFILIADO"
                texto[i]="Revisar datos del afiliado, mail no enviado. Entrega: " + str(entregas[i])
                resultado[i]="Enviado con error a Adm de Ventas"
                
                try:
                    
                    ######TRY EN CASO DE ERROR EN EL MAIL PARA ADM DE VENTAS#######
                    lector(str(entregas[i]),rutahtml,dia,farmacia[i])
                    enviarmail_sin_error("Scienza1","serviciopostentrega@scienza.com.ar",mail[i],asunto[i],texto[i],rutahtml,str(entregas[i])+".html","fcores@scienza.com.ar")
                
                except:
                    
                    ######TRY EN CASO DE ERROR EN LA IMPRESORA Y EN LOS DATOS DEL AFILIADO PARA ADM DE VENTAS#######
                    texto[i]="Revisar datos, el ticket no se encuentra en la impresora y los datos del afiliados son erroneos. Entrega: " + str(entregas[i])
                    resultado[i]="NOOK - Enviado con error a Adm de Ventas"
                    enviarmail_con_error("Scienza1","serviciopostentrega@scienza.com.ar",mail[i],asunto[i],texto[i],str(entregas[i])+".html","fcores@scienza.com.ar")
            
            else:
                try:
                    
                    ######TRY OK PARA ADM DE VENTAS#######
                    lector(str(entregas[i]),rutahtml,dia,str(farmacia[i]))
                    enviarmail_sin_error("Scienza1","serviciopostentrega@scienza.com.ar",mail[i],asunto[i],texto[i],rutahtml,str(entregas[i])+".html","fcores@scienza.com.ar")
                    resultado[i]="OK - Ticket enviado"
                except:
                    ######TRY EN CASO DE ERROR EN LA IMPRESORA PARA ADM DE VENTAS#######
                    resultado[i]="NOOK - Ticket no enviado, problemas en los datos de la operacion"
                    
            ws.cell(row=fila[i],column=8).value=resultado[i]
        wb.save("C:/Users/"+ user + "/Desktop/doka/doka_oyp.xlsm")
        wb.close()
    
    botonmail=Button(miFrame4, text="Mail",command=mail)
    botonmail.grid(row=2,column=3,sticky="e",padx=10,pady=10)

    botonarmado=Button(miFrame3, text="Inicio",command=cargardatos)
    botonarmado.grid(row=5,column=2,sticky="e",padx=10,pady=10)


    root.mainloop()
    
#---------------------------------------------------------fin-----------------------------------------------------------#
if __name__=="__main__":
    doka_oyp()   